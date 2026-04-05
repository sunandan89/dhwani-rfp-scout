#!/usr/bin/env python3
"""
rfp_scout.py  —  Dhwani RIS RFP Scouting Agent
================================================
Responsibility: Scrape RFP listings, score each for Dhwani relevance,
save results to JSON, and email a digest of shortlisted leads.

Does NOT draft proposals. To draft proposals, run proposal_agent.py
pointing it at the rfp_results.json this script produces.

Usage:
    python rfp_scout.py                   # Normal daily run
    python rfp_scout.py --no-email        # Skip sending email
    python rfp_scout.py --test            # Process first 3 RFPs only

Output (written to rfp_output/YYYY-MM-DD/):
    rfp_results.json         — all scored RFPs (input for proposal_agent.py)
    rfp_run_log.json         — run stats (for GitHub Actions summary)
    rfp_digest_*.html        — HTML email digest saved locally
    rfp_results_*.xlsx       — Excel report attached to email

Requirements:
    pip install requests beautifulsoup4 pyyaml anthropic openpyxl
"""

import os
import re
import json
import yaml
import time
import smtplib
import logging
import argparse
import requests
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from pathlib import Path

try:
    import anthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    ANTHROPIC_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from dhwani_profile import (
    DHWANI_PROFILE,
    RELEVANT_KEYWORDS,
    IRRELEVANT_KEYWORDS,
)


# ─────────────────────────────────────────────
#  SCRAPING
# ─────────────────────────────────────────────

def scrape_devnetjobsindia(test_mode=False):
    """Scrape RFP listings from DevNetJobsIndia."""
    url = "https://www.devnetjobsindia.org/rfp_assignments.aspx"
    headers = {
        'User-Agent': (
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
            'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        )
    }

    from bs4 import BeautifulSoup

    session = requests.Session()
    try:
        r = session.get(url, headers=headers, timeout=25)
        r.raise_for_status()
    except requests.RequestException as e:
        logging.error(f"Failed to fetch DevNetJobsIndia: {e}")
        return []

    soup = BeautifulSoup(r.text, 'html.parser')

    # Extract ASP.NET ViewState for postback simulation
    viewstate_el = soup.find('input', {'id': '__VIEWSTATE'})
    viewstate = viewstate_el['value'] if viewstate_el else ''
    viewstategen_el = soup.find('input', {'id': '__VIEWSTATEGENERATOR'})
    viewstategen = viewstategen_el['value'] if viewstategen_el else ''

    all_rows = soup.find_all('tr')
    grid_rows = [r for r in all_rows if r.find('a', href=lambda x: x and 'doPostBack' in str(x))]
    grid_rows = [r for r in grid_rows if r.find('span', id=lambda x: x and 'lblJobTitle' in str(x))]

    if test_mode:
        grid_rows = grid_rows[:3]

    rfps = []
    for i, row in enumerate(grid_rows):
        title_el  = row.find('span', id=lambda x: x and 'lblJobTitle' in str(x))
        org_el    = row.find('span', id=lambda x: x and 'lblJobCo' in str(x))
        loc_el    = row.find('span', id=lambda x: x and 'lblLocation' in str(x))
        dead_el   = row.find('span', id=lambda x: x and 'lblDeadline' in str(x))
        sector_el = row.find('span', id=lambda x: x and 'lblSectors' in str(x))

        title    = title_el.get_text(strip=True)  if title_el  else ''
        org      = org_el.get_text(strip=True)    if org_el    else ''
        location = loc_el.get_text(strip=True)    if loc_el    else ''
        deadline = dead_el.get_text(strip=True)   if dead_el   else ''
        sector   = sector_el.get_text(strip=True) if sector_el else ''

        if not title:
            continue

        # Try to extract job_id from logo image src (joblogos/XXXXX.png)
        img_el = row.find('img', src=lambda x: x and 'joblogos' in str(x))
        job_id = None
        if img_el:
            match = re.search(r'joblogos/(\d+)', img_el.get('src', ''))
            if match:
                job_id = match.group(1)

        # Postback target for fallback
        link_el = row.find('a', href=lambda x: x and 'doPostBack' in str(x) and 'lnkJobTitle' in str(x))
        postback_target = ''
        if link_el:
            match = re.search(r"doPostBack\('([^']+)'", link_el.get('href', ''))
            if match:
                postback_target = match.group(1)

        rfps.append({
            'title': title,
            'organization': org,
            'location': location,
            'deadline': deadline,
            'sector': sector,
            'job_id': job_id,
            'postback_target': postback_target,
            'source': 'DevNetJobsIndia',
            'row_index': i,
            'full_description': None,
            'url': None,
        })

    logging.info(f"DevNetJobsIndia: found {len(rfps)} RFP listings")

    # Fetch full descriptions
    post_headers = {**headers,
                    'Content-Type': 'application/x-www-form-urlencoded',
                    'Referer': url}

    for rfp in rfps:
        try:
            if rfp['job_id']:
                detail_url = f"https://devnetjobsindia.org/JobDescription.aspx?Job_Id={rfp['job_id']}"
                resp = session.get(detail_url, headers=headers, timeout=20)
                rfp['url'] = detail_url
                rfp['full_description'] = _extract_devnet_description(resp.text)
            elif rfp['postback_target']:
                post_data = {
                    '__EVENTTARGET':        rfp['postback_target'],
                    '__EVENTARGUMENT':      '',
                    '__VIEWSTATE':          viewstate,
                    '__VIEWSTATEGENERATOR': viewstategen,
                }
                resp = session.post(url, data=post_data, headers=post_headers,
                                    timeout=20, allow_redirects=True)
                match = re.search(r'Job_Id=(\d+)', resp.url, re.IGNORECASE)
                if match:
                    rfp['job_id'] = match.group(1)
                    rfp['url'] = resp.url
                rfp['full_description'] = _extract_devnet_description(resp.text)
                time.sleep(0.6)
        except Exception as e:
            logging.warning(f"Could not fetch detail for '{rfp['title'][:40]}': {e}")

    return rfps


def _extract_devnet_description(html):
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, 'html.parser')
    for tag in soup.find_all(['nav', 'header', 'footer', 'script', 'style', 'noscript']):
        tag.decompose()
    main = soup.find('div', id=lambda x: x and 'ContentPlaceHolder1' in str(x))
    if main:
        text = main.get_text(separator='\n', strip=True)
        text = re.sub(r'\n{3,}', '\n\n', text)
        return text[:4000]
    return soup.get_text(separator='\n', strip=True)[:3000]


def try_scrape_devex(config):
    """
    Optional: scrape Devex. Requires a session cookie from a logged-in browser session.
    Set devex_cookie in config.yaml or DEVEX_COOKIE env var to enable.
    """
    from bs4 import BeautifulSoup
    rfps = []
    cookie = os.environ.get('DEVEX_COOKIE') or config.get('devex_cookie', '')
    if not cookie:
        logging.info("Devex: No session cookie configured — skipping. "
                     "(Add devex_cookie to config.yaml or set DEVEX_COOKIE env var)")
        return rfps

    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        'Cookie': cookie,
    }
    try:
        r = requests.get('https://www.devex.com/jobs/rfps', headers=headers, timeout=20)
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')
            cards = soup.find_all('div', class_=lambda x: x and 'job' in str(x).lower())
            for card in cards[:20]:
                title_el = card.find(['h2', 'h3', 'a'])
                title = title_el.get_text(strip=True) if title_el else ''
                if title:
                    rfps.append({
                        'title': title, 'organization': '', 'location': '',
                        'deadline': '', 'sector': '', 'source': 'Devex',
                        'url': 'https://www.devex.com/jobs/rfps',
                        'full_description': None, 'job_id': None,
                    })
    except Exception as e:
        logging.warning(f"Devex scraping failed: {e}")

    logging.info(f"Devex: found {len(rfps)} RFPs")
    return rfps


# ─────────────────────────────────────────────
#  SCORING
# ─────────────────────────────────────────────

def score_rfp(rfp, claude_client):
    """Score RFP relevance. Uses Claude if available, else keyword fallback."""
    if claude_client:
        return _score_with_claude(rfp, claude_client)
    return _score_with_keywords(rfp)


def _score_with_claude(rfp, client):
    desc = (rfp.get('full_description') or '')[:2000]
    prompt = f"""You are a business development expert for Dhwani RIS — an Indian IT company serving the social sector.

DHWANI PROFILE:
{DHWANI_PROFILE}

EVALUATE THIS RFP:
Title: {rfp['title']}
Organization: {rfp['organization']}
Sector: {rfp.get('sector', '')}
Location: {rfp.get('location', '')}
Deadline: {rfp.get('deadline', '')}
Description: {desc if desc else 'Not available'}

Score 0–10 for Dhwani RIS relevance:
10 = Perfect (grant management software, MIS platform, data collection app, HMIS)
7–9 = Good fit (tech solution in a sector Dhwani serves)
4–6 = Moderate (some tech component but not Dhwani's core)
1–3 = Weak (minimal tech)
0 = Not relevant (physical goods, housekeeping, travel, CA audit, film production, etc.)

Respond ONLY with valid JSON:
{{
  "score": <0-10 integer>,
  "recommendation": "apply" | "consider" | "skip",
  "reason": "<one concise sentence>",
  "relevant_product": "<mGrant | mForm | mLearn | Custom Dev | Tech Consulting | DaaS | None>",
  "key_requirements": ["<req1>", "<req2>", "<req3>"]
}}"""

    try:
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=400,
            messages=[{"role": "user", "content": prompt}]
        )
        text = msg.content[0].text.strip()
        json_match = re.search(r'\{.*\}', text, re.DOTALL)
        if json_match:
            return json.loads(json_match.group())
    except Exception as e:
        logging.warning(f"Claude scoring failed for '{rfp['title'][:40]}': {e}")

    return _score_with_keywords(rfp)


def _score_with_keywords(rfp):
    text = ' '.join([
        rfp.get('title', ''), rfp.get('sector', ''),
        rfp.get('organization', ''), rfp.get('full_description', '') or ''
    ]).lower()

    score = sum(1 for kw in RELEVANT_KEYWORDS if kw.lower() in text)
    score -= sum(3 for kw in IRRELEVANT_KEYWORDS if kw.lower() in text)
    score = max(0, min(10, score))
    rec = 'apply' if score >= 7 else ('consider' if score >= 4 else 'skip')

    return {
        'score': score,
        'recommendation': rec,
        'reason': 'Based on keyword matching (Claude API not configured)',
        'relevant_product': 'Unknown',
        'key_requirements': [],
    }


# ─────────────────────────────────────────────
#  EXCEL REPORT
# ─────────────────────────────────────────────

def generate_excel_report(all_rfps, output_dir):
    """
    Generate a branded Excel report with two sheets:
      • Shortlisted  — RFPs recommended to apply or consider
      • All RFPs     — Everything scraped with scores
    Returns path to the saved .xlsx file, or None if openpyxl is unavailable.
    """
    if not OPENPYXL_AVAILABLE:
        logging.warning("openpyxl not installed — skipping Excel report. pip install openpyxl")
        return None

    today = datetime.now().strftime('%Y-%m-%d')
    filename = f"rfp_results_{today}.xlsx"
    filepath = Path(output_dir) / filename

    wb = Workbook()

    # ── Colour palette ──────────────────────────────────────
    RED_DARK   = "8B0000"
    RED        = "C00000"
    WHITE      = "FFFFFF"
    GREEN_BG   = "E8F5E9"
    GREEN_FONT = "1B5E20"
    AMBER_BG   = "FFF8E1"
    AMBER_FONT = "E65100"
    GREY_BG    = "F5F5F5"
    HEADER_BG  = "C00000"

    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )

    def _style_header_row(ws, row_num, columns):
        for col_num, header in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = Font(name='Arial', bold=True, color=WHITE, size=10)
            cell.fill = PatternFill('solid', fgColor=HEADER_BG)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    def _style_data_row(ws, row_num, values, bg_color):
        for col_num, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=val)
            cell.font = Font(name='Arial', size=9)
            cell.fill = PatternFill('solid', fgColor=bg_color)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border

    # ═══════════════════════════════════════════════════════
    #  SHEET 1 — Shortlisted RFPs
    # ═══════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Shortlisted RFPs"

    # Title banner
    ws1.merge_cells('A1:J1')
    banner = ws1['A1']
    banner.value = f"Dhwani RIS — RFP Scout Results  |  {datetime.now().strftime('%d %B %Y')}"
    banner.font = Font(name='Arial', bold=True, size=14, color=WHITE)
    banner.fill = PatternFill('solid', fgColor=RED_DARK)
    banner.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 30

    # Sub-banner
    ws1.merge_cells('A2:J2')
    shortlisted = [r for r in all_rfps
                   if r.get('scoring', {}).get('recommendation') in ('apply', 'consider')]
    apply_n   = sum(1 for r in shortlisted if r.get('scoring', {}).get('recommendation') == 'apply')
    consider_n = len(shortlisted) - apply_n
    sub = ws1['A2']
    sub.value = (f"✅ {apply_n} to Apply   🤔 {consider_n} to Consider   "
                 f"📋 {len(all_rfps)} total scraped")
    sub.font = Font(name='Arial', size=10, bold=True, color=RED)
    sub.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[2].height = 20

    # Empty row
    ws1.row_dimensions[3].height = 6

    # Headers
    COLS_SHORT = [
        "Rank", "Score\n(/10)", "Recommendation", "Title",
        "Organization", "Location", "Deadline", "Sector",
        "Relevant\nProduct", "URL"
    ]
    _style_header_row(ws1, 4, COLS_SHORT)
    ws1.row_dimensions[4].height = 32

    # Data rows
    rank = 0
    for rfp in sorted(shortlisted,
                      key=lambda r: r.get('scoring', {}).get('score', 0), reverse=True):
        sc  = rfp.get('scoring', {})
        rec = sc.get('recommendation', 'consider')
        rank += 1

        bg = GREEN_BG if rec == 'apply' else AMBER_BG
        rec_label = "✅ APPLY" if rec == 'apply' else "🤔 CONSIDER"

        row_vals = [
            rank,
            sc.get('score', 0),
            rec_label,
            rfp.get('title', ''),
            rfp.get('organization', ''),
            rfp.get('location', ''),
            rfp.get('deadline', ''),
            rfp.get('sector', ''),
            sc.get('relevant_product', ''),
            rfp.get('url', ''),
        ]
        row_num = rank + 4
        _style_data_row(ws1, row_num, row_vals, bg)
        ws1.row_dimensions[row_num].height = 36

        # Make score bold
        score_cell = ws1.cell(row=row_num, column=2)
        score_cell.font = Font(name='Arial', size=10, bold=True,
                               color=GREEN_FONT if rec == 'apply' else AMBER_FONT)
        score_cell.alignment = Alignment(horizontal='center', vertical='top')

        # Make rec bold
        rec_cell = ws1.cell(row=row_num, column=3)
        rec_cell.font = Font(name='Arial', size=9, bold=True,
                              color=GREEN_FONT if rec == 'apply' else AMBER_FONT)
        rec_cell.alignment = Alignment(horizontal='center', vertical='top')

        # URL as hyperlink
        url = rfp.get('url', '')
        if url:
            url_cell = ws1.cell(row=row_num, column=10)
            url_cell.value = "View RFP →"
            url_cell.hyperlink = url
            url_cell.font = Font(name='Arial', size=9, color="0066CC",
                                  underline='single')

    # Column widths (Sheet 1)
    col_widths_s1 = [6, 8, 14, 48, 30, 18, 16, 20, 18, 14]
    for i, w in enumerate(col_widths_s1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes below header
    ws1.freeze_panes = 'A5'

    # ═══════════════════════════════════════════════════════
    #  SHEET 2 — All Scraped RFPs
    # ═══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("All RFPs")

    ws2.merge_cells('A1:I1')
    b2 = ws2['A1']
    b2.value = f"All Scraped RFPs — {datetime.now().strftime('%d %B %Y')}  ({len(all_rfps)} total)"
    b2.font = Font(name='Arial', bold=True, size=12, color=WHITE)
    b2.fill = PatternFill('solid', fgColor=RED_DARK)
    b2.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 26

    COLS_ALL = [
        "Score\n(/10)", "Recommendation", "Title",
        "Organization", "Location", "Deadline", "Sector",
        "Reason / Scoring Note", "URL"
    ]
    _style_header_row(ws2, 2, COLS_ALL)
    ws2.row_dimensions[2].height = 32

    for row_idx, rfp in enumerate(
            sorted(all_rfps, key=lambda r: r.get('scoring', {}).get('score', 0), reverse=True),
            start=3):
        sc  = rfp.get('scoring', {})
        rec = sc.get('recommendation', 'skip')
        bg = (GREEN_BG if rec == 'apply'
              else (AMBER_BG if rec == 'consider' else GREY_BG))

        row_vals = [
            sc.get('score', 0),
            rec.upper(),
            rfp.get('title', ''),
            rfp.get('organization', ''),
            rfp.get('location', ''),
            rfp.get('deadline', ''),
            rfp.get('sector', ''),
            sc.get('reason', ''),
            rfp.get('url', ''),
        ]
        _style_data_row(ws2, row_idx, row_vals, bg)
        ws2.row_dimensions[row_idx].height = 28

        # Hyperlink URL
        url = rfp.get('url', '')
        if url:
            url_cell = ws2.cell(row=row_idx, column=9)
            url_cell.value = "View →"
            url_cell.hyperlink = url
            url_cell.font = Font(name='Arial', size=9, color="0066CC", underline='single')

    col_widths_s2 = [8, 14, 48, 30, 18, 16, 20, 45, 10]
    for i, w in enumerate(col_widths_s2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.freeze_panes = 'A3'

    wb.save(str(filepath))
    logging.info(f"Excel report saved: {filepath}")
    return str(filepath)


# ─────────────────────────────────────────────
#  EMAIL DIGEST
# ─────────────────────────────────────────────

def send_email_digest(relevant_rfps, config, output_dir, excel_file=None):
    """
    Build and send the daily digest email.
    Attaches the Excel report (rfp_results_*.xlsx) when excel_file is provided.
    NOTE: Proposals are NOT attached here — that is proposal_agent.py's job.
    """
    today_str = datetime.now().strftime('%d %B %Y')
    apply_count   = sum(1 for r in relevant_rfps
                        if r.get('scoring', {}).get('recommendation') == 'apply')
    consider_count = sum(1 for r in relevant_rfps
                         if r.get('scoring', {}).get('recommendation') == 'consider')

    excel_note = (
        '<div style="background:#E8F5E9; border-left:4px solid #2e7d32; padding:10px 16px; '
        'margin-bottom:16px; border-radius:0 6px 6px 0; font-size:13px;">'
        '📎 <strong>Full results attached as Excel spreadsheet</strong> '
        '(rfp_results_' + datetime.now().strftime('%Y-%m-%d') + '.xlsx)'
        '</div>'
    ) if excel_file else ''

    html = f"""<!DOCTYPE html>
<html>
<body style="font-family: Arial, sans-serif; max-width: 750px; margin: 0 auto; padding: 20px; color: #333;">

<div style="background: linear-gradient(135deg, #c00000, #8b0000); padding: 20px 25px; border-radius: 8px; margin-bottom: 20px;">
  <h1 style="color: white; margin: 0; font-size: 22px;">🎯 Daily RFP Scout</h1>
  <p style="color: #ffcccc; margin: 6px 0 0 0; font-size: 14px;">Dhwani RIS Business Development &nbsp;|&nbsp; {today_str}</p>
</div>

<div style="background: #f7f7f7; border-left: 4px solid #c00000; padding: 12px 16px; margin-bottom: 16px; border-radius: 0 6px 6px 0;">
  <strong>Today's Summary:</strong>
  &nbsp; ✅ <strong>{apply_count}</strong> to Apply
  &nbsp; 🤔 <strong>{consider_count}</strong> to Consider
  <br><small style="color: #888; margin-top:4px; display:block;">
    To generate proposal drafts, run: <code>python proposal_agent.py --from-results rfp_output/YYYY-MM-DD/rfp_results.json</code>
  </small>
</div>
{excel_note}
"""

    for rfp in relevant_rfps:
        sc     = rfp.get('scoring', {})
        score  = sc.get('score', 0)
        rec    = sc.get('recommendation', 'consider')
        reason = sc.get('reason', '')
        product = sc.get('relevant_product', '')

        color = '#1a7a1a' if score >= 8 else ('#e07800' if score >= 5 else '#c00000')
        badge = {'apply': '✅ APPLY NOW', 'consider': '🤔 CONSIDER', 'skip': '⏭ SKIP'}.get(rec, rec.upper())

        html += f"""
<div style="border: 1px solid #e0e0e0; border-radius: 8px; padding: 16px; margin-bottom: 16px; box-shadow: 0 1px 3px rgba(0,0,0,0.07);">
  <div style="display:flex; justify-content:space-between; align-items:flex-start; flex-wrap:wrap; gap:8px;">
    <h3 style="margin:0; font-size:16px; color:#1a1a1a; flex:1;">{rfp['title']}</h3>
    <span style="background:{color}; color:white; padding:4px 12px; border-radius:20px; font-size:13px; font-weight:bold; white-space:nowrap;">{score}/10</span>
  </div>
  <table style="width:100%; margin-top:10px; font-size:13px; border-collapse:collapse;">
    <tr><td style="color:#666; width:130px; padding:3px 0;">Organization</td><td><strong>{rfp['organization']}</strong></td></tr>
    <tr><td style="color:#666; padding:3px 0;">Location</td><td>{rfp.get('location','N/A')}</td></tr>
    <tr><td style="color:#666; padding:3px 0;">Deadline</td><td><strong style="color:#c00000;">{rfp.get('deadline','N/A')}</strong></td></tr>
    <tr><td style="color:#666; padding:3px 0;">Source</td><td>{rfp.get('source','N/A')} &nbsp; <a href="{rfp.get('url','#')}" style="color:#0066cc;">View RFP →</a></td></tr>
    <tr><td style="color:#666; padding:3px 0;">Relevant Product</td><td>{product}</td></tr>
  </table>
  <div style="margin-top:10px; padding:8px 12px; background:#f9f9f9; border-radius:5px; font-size:13px;">
    <strong>{badge}</strong> — {reason}
  </div>
</div>
"""

    html += f"""
<div style="margin-top:24px; padding:14px; background:#f0f0f0; border-radius:6px; font-size:11px; color:#888; text-align:center;">
  Automated daily digest by <strong>Dhwani RIS RFP Scout</strong><br>
  To draft proposals, run <code>proposal_agent.py</code> on today's results.<br>
  Contact: reachus@dhwaniris.com | www.dhwaniris.in
</div>
</body></html>"""

    # Save HTML locally
    digest_path = Path(output_dir) / f"rfp_digest_{datetime.now().strftime('%Y-%m-%d')}.html"
    with open(digest_path, 'w', encoding='utf-8') as f:
        f.write(html)
    logging.info(f"HTML digest saved: {digest_path}")

    # Override config from environment variables (GitHub Actions / CI)
    email_cfg = config.get('email', {})
    env_sender     = os.environ.get('EMAIL_SENDER')
    env_password   = os.environ.get('EMAIL_PASSWORD')
    env_recipients = os.environ.get('EMAIL_RECIPIENTS')
    env_smtp_host  = os.environ.get('EMAIL_SMTP_HOST')
    env_smtp_port  = os.environ.get('EMAIL_SMTP_PORT')

    if env_sender:     email_cfg['sender_email'] = env_sender
    if env_password:   email_cfg['password']     = env_password
    if env_recipients: email_cfg['recipients']   = [r.strip() for r in env_recipients.split(',')]
    if env_smtp_host:  email_cfg['smtp_host']    = env_smtp_host
    if env_smtp_port:  email_cfg['smtp_port']    = int(env_smtp_port)
    if env_sender and env_password and env_recipients:
        email_cfg['enabled'] = True

    if not email_cfg.get('enabled'):
        logging.info("Email not enabled. Digest saved to output folder only.")
        return str(digest_path)

    msg = MIMEMultipart('mixed')
    msg['Subject'] = f"[RFP Scout] {apply_count} to Apply, {consider_count} to Consider — {today_str}"
    msg['From']    = email_cfg['sender_email']
    msg['To']      = ', '.join(email_cfg['recipients'])
    msg.attach(MIMEText(html, 'html'))

    # Attach Excel report
    if excel_file and Path(excel_file).exists():
        with open(excel_file, 'rb') as f:
            part = MIMEApplication(
                f.read(),
                _subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            part.add_header('Content-Disposition', 'attachment',
                             filename=Path(excel_file).name)
            msg.attach(part)
        logging.info(f"  → Attaching Excel: {Path(excel_file).name}")

    try:
        smtp_host = email_cfg.get('smtp_host', 'smtp.office365.com')
        smtp_port = email_cfg.get('smtp_port', 587)
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.ehlo()
            server.starttls()
            server.login(email_cfg['sender_email'], email_cfg['password'])
            server.sendmail(email_cfg['sender_email'], email_cfg['recipients'], msg.as_string())
        logging.info(f"✉ Email sent to: {email_cfg['recipients']}")
    except Exception as e:
        logging.error(f"Email send failed: {e}")

    return str(digest_path)


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Dhwani RIS RFP Scout')
    parser.add_argument('--no-email',  action='store_true', help='Skip email')
    parser.add_argument('--test',      action='store_true', help='Process first 3 RFPs only')
    parser.add_argument('--min-score', type=int, default=None, help='Min score to include (default from config: 5)')
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        handlers=[
            logging.FileHandler('rfp_scout.log', encoding='utf-8'),
            logging.StreamHandler()
        ]
    )

    # Load config
    config_path = Path(__file__).parent / 'config.yaml'
    config = {}
    if config_path.exists():
        with open(config_path) as f:
            config = yaml.safe_load(f) or {}

    if args.no_email:
        config.setdefault('email', {})['enabled'] = False

    # Output directory
    base_out = Path(config.get('output_dir', './rfp_output'))
    today_out = base_out / datetime.now().strftime('%Y-%m-%d')
    today_out.mkdir(parents=True, exist_ok=True)

    # Claude client (optional — for smarter scoring)
    api_key = config.get('anthropic_api_key') or os.environ.get('ANTHROPIC_API_KEY')
    claude_client = None
    if ANTHROPIC_AVAILABLE and api_key and 'YOUR_ANTHROPIC' not in api_key:
        claude_client = anthropic.Anthropic(api_key=api_key)
        logging.info("Claude API: connected ✓")
    else:
        logging.warning("Claude API not configured — using keyword scoring.")

    logging.info("=" * 65)
    logging.info(f"  Dhwani RIS RFP Scout  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    if args.test:
        logging.info("  MODE: TEST (first 3 RFPs only)")
    logging.info("=" * 65)

    # ── Scrape ────────────────────────────────
    all_rfps = []
    logging.info("▶ Scraping DevNetJobsIndia...")
    devnet = scrape_devnetjobsindia(test_mode=args.test)
    all_rfps.extend(devnet)
    logging.info(f"  → {len(devnet)} RFPs fetched")

    logging.info("▶ Trying Devex...")
    devex = try_scrape_devex(config.get('devex', {}))
    all_rfps.extend(devex)

    logging.info(f"Total RFPs to evaluate: {len(all_rfps)}")

    # ── Score ─────────────────────────────────
    logging.info("▶ Scoring RFPs for Dhwani relevance...")
    min_score = args.min_score if args.min_score is not None else config.get('min_score', 5)
    relevant = []

    for rfp in all_rfps:
        logging.info(f"  Evaluating: {rfp['title'][:60]}")
        scoring = score_rfp(rfp, claude_client)
        rfp['scoring'] = scoring
        score = scoring.get('score', 0)
        rec   = scoring.get('recommendation', 'skip')
        logging.info(f"    Score: {score}/10  |  {rec.upper()}  |  {scoring.get('reason','')}")

        if rec in ('apply', 'consider') and score >= min_score:
            relevant.append(rfp)

        time.sleep(0.3)

    logging.info(f"\n✓ {len(relevant)} relevant RFP(s) found (score ≥ {min_score})")

    # ── Save results JSON (handoff to proposal_agent.py) ─────
    results_file = today_out / 'rfp_results.json'
    with open(results_file, 'w', encoding='utf-8') as f:
        json.dump(all_rfps, f, indent=2, ensure_ascii=False)
    logging.info(f"Results saved: {results_file}")
    logging.info(f"  → To draft proposals: python proposal_agent.py --from-results {results_file}")

    # ── Generate Excel report ─────────────────────────────
    logging.info("▶ Generating Excel report...")
    excel_file = generate_excel_report(all_rfps, str(today_out))

    # ── Save run log (for GitHub Actions summary) ─────────────
    log_data = {
        'run_date': datetime.now().isoformat(),
        'total_scraped': len(all_rfps),
        'relevant_count': len(relevant),
        'proposals_drafted': 0,  # proposal_agent handles this
        'rfps': [{
            'title':          r['title'],
            'organization':   r['organization'],
            'deadline':       r.get('deadline'),
            'score':          r.get('scoring', {}).get('score'),
            'recommendation': r.get('scoring', {}).get('recommendation'),
            'reason':         r.get('scoring', {}).get('reason'),
            'url':            r.get('url'),
            'has_proposal':   False,
        } for r in all_rfps]
    }
    log_file = today_out / 'rfp_run_log.json'
    with open(log_file, 'w', encoding='utf-8') as f:
        json.dump(log_data, f, indent=2, ensure_ascii=False)

    # ── Email digest (with Excel attached) ───────────────────
    if relevant:
        logging.info("▶ Sending email digest with Excel attachment...")
        send_email_digest(relevant, config, str(today_out), excel_file=excel_file)
    else:
        logging.info("No relevant RFPs today — no digest sent.")

    # ── Summary ───────────────────────────────
    logging.info("\n" + "=" * 65)
    logging.info("  SCOUT RUN COMPLETE")
    logging.info(f"  Total scraped  : {len(all_rfps)}")
    logging.info(f"  Relevant       : {len(relevant)}")
    logging.info(f"  Output folder  : {today_out}")
    logging.info(f"  Results JSON   : {results_file}")
    if excel_file:
        logging.info(f"  Excel report   : {excel_file}")
    logging.info("=" * 65)

    return relevant


if __name__ == '__main__':
    main()
